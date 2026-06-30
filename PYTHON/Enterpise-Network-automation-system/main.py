from netmiko import ConnectHandler
import openpyxl
import os
import time

EXCEL_FILE = "tabel1.xlsx"
SSH_USER = "admin"
SSH_PASS = "key123"

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"[!] File {EXCEL_FILE} kagak ada jings!")
        return

    print(f"[*] Membuka dan membaca data dari {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    network_data = []
    for row_idx in range(4, sheet.max_row + 1):
        router_name = sheet.cell(row=row_idx, column=2).value
        ip_ssh = sheet.cell(row=row_idx, column=3).value
        vlan_id = sheet.cell(row=row_idx, column=4).value
        vlan_name = sheet.cell(row=row_idx, column=5).value
        interface = sheet.cell(row=row_idx, column=6).value
        gateway = sheet.cell(row=row_idx, column=7).value
        pool_raw = sheet.cell(row=row_idx, column=8).value
        
        if not router_name:
            continue
            
        pool = str(pool_raw).replace(',', '.')
        ip_base = gateway.split('/')[0]
        network_prefix = '.'.join(ip_base.split('.')[:3]) + ".0/" + gateway.split('/')[1]
        gateway_ip = ip_base

        network_data.append({
            "row_idx": row_idx,
            "router_name": router_name,
            "ip": ip_ssh,
            "vlan_id": vlan_id,
            "vlan_name": vlan_name,
            "interface": interface,
            "gateway": gateway,
            "pool": pool,
            "network": network_prefix,
            "gateway_ip": gateway_ip
        })

    devices = {}
    for data in network_data:
        ip = data["ip"]
        if ip not in devices:
            devices[ip] = {"name": data["router_name"], "configs": []}
        devices[ip]["configs"].append(data)

    for ip, dev_info in devices.items():
        print(f"\n==================== PROSES ROUTER: {dev_info['name']} ({ip}) ====================")
        
        mikrotik_device = {
            'device_type': 'mikrotik_routeros',
            'host': ip,
            'username': SSH_USER,
            'password': SSH_PASS,
            'port': 22,
        }
        
        try:
            print(f"[*] Menghubungi router {dev_info['name']} via SSH...")
            net_connect = ConnectHandler(**mikrotik_device)
            
            success_rows = []
            for config in dev_info["configs"]:
                v_id = config["vlan_id"]
                v_name = config["vlan_name"]  # VLAN-10, dst.
                iface = config["interface"]   # ether port fisik target
                gw = config["gateway"]
                pool = config["pool"]
                net_add = config["network"]
                gw_ip = config["gateway_ip"]
                
                bridge_name = f"bridge-{v_name}" # Bikin nama bridge terpisah per VLAN!

                print(f"\n    [>] CONFIG ALL-IN-ONE {v_name} PADA {bridge_name}:")
                
                # 1. BIKIN INTERFACE BRIDGE NYA SENDIRI-SENDIRI (BANYAK BRIDGE)
                print(f"        -> Membuat {bridge_name}...")
                net_connect.send_command(f"/interface bridge add name={bridge_name} comment=\"Bridge khusus {v_name}\"")
                
                # 2. MASUKIN PORT FISIK KE BRIDGE MASING-MASING
                print(f"        -> Memasukkan port {iface} ke {bridge_name}...")
                net_connect.send_command(f"/interface bridge port add bridge={bridge_name} interface={iface}")
                
                # 3. BIKIN INTERFACE VLAN RESMI DI ATAS BRIDGE MASING-MASING
                print(f"        -> Membuat Interface VLAN resmi {v_name} (ID: {v_id}) di atas {bridge_name}...")
                net_connect.send_command(f"/interface vlan add name={v_name} vlan-id={v_id} interface={bridge_name}")
                
                # 4. PASANG IP ADDRESS/GATEWAY DI INTERFACE VLAN-NYA
                print(f"        -> Menempelkan IP Gateway {gw} ke interface {v_name}...")
                net_connect.send_command(f"/ip address add address={gw} interface={v_name} comment=\"{v_name}\"")
                
                # 5. SETUP FULL DHCP SERVER DI ATAS INTERFACE VLAN KELUARAN BRIDGE
                print(f"        -> Konfigurasi Pool, DHCP Server & Network untuk {v_name}...")
                net_connect.send_command(f"/ip pool add name=pool-{v_name} ranges={pool}")
                net_connect.send_command(f"/ip dhcp-server add name=dhcp-{v_name} interface={v_name} pool=pool-{v_name} disabled=no")
                net_connect.send_command(f"/ip dhcp-server network add address={net_add} gateway={gw_ip} dns-server=8.8.8.8")
                
                success_rows.append(config["row_idx"])
            
            print(f"\n[+] {dev_info['name']} BERHASIL DISUNTIK SKEMA MULTI-BRIDGE LENGKAP!")
            for r_idx in success_rows:
                sheet.cell(row=r_idx, column=9).value = "SUCCESS"
                
            net_connect.disconnect()
            
        except Exception as e:
            print(f"[!] Gagal eksekusi massal pada {dev_info['name']}: {e}")
            for config in dev_info["configs"]:
                sheet.cell(row=config["row_idx"], column=9).value = "FAILED"

    print(f"\n[*] Menulis status terbaru ke dalam file {EXCEL_FILE}...")
    wb.save(EXCEL_FILE)
    print("[+] Selesai jings! Excel ter-update otomatis jadi SUCCESS!")

if __name__ == "__main__":
    main()